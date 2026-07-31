"""
census_extractor.py
--------------------
Universal reader for Census Excel workbooks. Auto-detects the header row
and maps columns using fuzzy matching against known field aliases defined
in config.CENSUS_HEADER_ALIASES.

Falls back to hardcoded positions (config.CENSUS_FIELD_COLS) if auto-detect
fails. Supports both .xlsx and .xls files.
"""
from __future__ import annotations

import difflib
import re
from dataclasses import dataclass
from datetime import datetime, date
from pathlib import Path

import openpyxl

from config import (
    CENSUS_FIELD_COLS,
    CENSUS_FIRST_DATA_ROW,
    CENSUS_HEADER_ROW,
    CENSUS_SHEET_NAME,
    CENSUS_HEADER_ALIASES,
)


@dataclass
class CensusRow:
    row_no: object
    first_name: str
    last_name: str
    gender: str | None
    dob: object
    home_zip: str | None
    relationship: str | None
    dependent_of_employee_number: object
    medical_coverage_tier: str | None
    cobra: str | None
    medical_plan_enrolled: str | None = None
    medical_plan_price: object = None
    dental_coverage_tier: str | None = None
    dental_plan_enrolled: str | None = None
    dental_plan_price: object = None
    vision_coverage_tier: str | None = None
    vision_plan_enrolled: str | None = None
    vision_plan_price: object = None
    life_plan_name: str | None = None
    life_benefit: object = None
    life_rate: object = None
    ltd_plan: str | None = None
    ltd_benefit: object = None
    ltd_rate: object = None
    std_plan: str | None = None
    std_benefit: object = None
    std_rate: object = None
    work_state: str | None = None
    job_title: str | None = None
    workers_comp_code: object = None
    annual_salary: object = None
    ft_pt: str | None = None

    @property
    def full_name_key(self) -> str:
        return f"{(self.first_name or '').strip().lower()} {(self.last_name or '').strip().lower()}"

    @property
    def is_dependent(self) -> bool:
        if not self.relationship:
            return False
        return self.relationship.strip().upper() not in ("EE", "EMPLOYEE")


def _clean(value):
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


# ---------------------------------------------------------------------------
# Auto-detect header row and column mapping
# ---------------------------------------------------------------------------

def _normalize_header(header: str) -> str:
    if header is None:
        return ""
    h = str(header)
    h = re.sub(r'\(.*?\)', '', h)
    return re.sub(r'[^a-zA-Z0-9]', '', h).lower()


def _find_best_field_match(header_text: str) -> str | None:
    """
    Given a header cell's text, find the best matching field name from
    CENSUS_HEADER_ALIASES. Returns the field name or None if no match.

    When a header matches aliases from multiple fields exactly (e.g. "Tier"
    matches both medical_coverage_tier and tier_generic), prefer the
    "_generic" field because the header is ambiguous and should be handled
    by the pivot normalizer rather than being assigned to a specific plan type.
    """
    normalized = _normalize_header(header_text)
    if not normalized:
        return None

    # Collect ALL exact matches, then pick the best one
    exact_matches = []
    best_field = None
    best_score = 0.0

    for field_name, aliases in CENSUS_HEADER_ALIASES.items():
        for alias in aliases:
            # Exact match (case-insensitive)
            if normalized == alias.lower():
                exact_matches.append(field_name)
                break  # Don't count multiple aliases from the same field

            # Check if alias is contained within the header text
            if alias.lower() in normalized:
                score = len(alias) / len(normalized)  # Prefer longer matches
                if score > best_score and score > 0.5:
                    best_score = score
                    best_field = field_name

    # Resolve exact matches: prefer _generic fields for ambiguous headers
    if exact_matches:
        # If there's a _generic version among the exact matches, use it
        generic = [f for f in exact_matches if f.endswith("_generic")]
        if generic:
            return generic[0]
        return exact_matches[0]

    if best_field is not None:
        return best_field

    # Fuzzy fallback for close matches
    all_aliases = []
    alias_to_field = {}
    for field_name, aliases in CENSUS_HEADER_ALIASES.items():
        for alias in aliases:
            all_aliases.append(alias.lower())
            alias_to_field[alias.lower()] = field_name
    
    close = difflib.get_close_matches(normalized, all_aliases, n=1, cutoff=0.75)
    if close:
        return alias_to_field[close[0]]

    return None


def _auto_detect_layout(ws, max_scan_rows: int = 50) -> tuple[int, dict[str, int]] | None:
    """
    Scan the first `max_scan_rows` rows of the worksheet looking for the
    header row. Returns (header_row, field_cols_dict) or None if detection fails.
    
    The header row is the one with the most cells that match known field aliases.
    """
    best_row = None
    best_score = 0
    best_mapping = {}

    max_col = min(ws.max_column or 50, 50)

    for row_idx in range(1, min(max_scan_rows + 1, (ws.max_row or max_scan_rows) + 1)):
        mapping = {}
        score = 0
        used_fields = set()

        for col_idx in range(1, max_col + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            field = _find_best_field_match(str(cell_value) if cell_value is not None else "")
            if field and field not in used_fields:
                mapping[field] = col_idx
                used_fields.add(field)
                score += 1

        # Require at least first_name and last_name to be found
        if score > best_score and "first_name" in mapping and "last_name" in mapping:
            best_score = score
            best_row = row_idx
            best_mapping = mapping

    if best_row is not None and best_score >= 3:
        return best_row, best_mapping
    
    return None


def _find_first_data_row(ws, header_row: int, field_cols: dict[str, int]) -> int:
    """Find the first row after the header that contains actual data."""
    first_name_col = field_cols.get("first_name")
    last_name_col = field_cols.get("last_name")
    
    for row_idx in range(header_row + 1, header_row + 10):
        first_val = ws.cell(row=row_idx, column=first_name_col).value if first_name_col else None
        last_val = ws.cell(row=row_idx, column=last_name_col).value if last_name_col else None
        
        if first_val is not None or last_val is not None:
            return row_idx
    
    # Default: row right after header
    return header_row + 1


def _find_sheet(wb) -> str:
    """Find the best sheet name to use."""
    # Try exact match first
    if CENSUS_SHEET_NAME in wb.sheetnames:
        return CENSUS_SHEET_NAME
    
    # Try case-insensitive match
    for name in wb.sheetnames:
        if name.lower() == CENSUS_SHEET_NAME.lower():
            return name
    
    # Try partial match
    for name in wb.sheetnames:
        if "census" in name.lower():
            return name
    
    # Fall back to first sheet
    return wb.sheetnames[0]


# ---------------------------------------------------------------------------
# Main extraction
# ---------------------------------------------------------------------------

def extract_census(xlsx_path: Path) -> list[CensusRow]:
    """Read every populated data row from Census Excel into CensusRow objects.
    
    Auto-detects header row and column layout. Falls back to hardcoded
    positions if auto-detection fails.
    """
    # Handle .xls and .csv files by converting first
    if xlsx_path.suffix.lower() == ".xls":
        xlsx_path = _convert_xls_to_xlsx(xlsx_path)
    elif xlsx_path.suffix.lower() == ".csv":
        xlsx_path = _convert_csv_to_xlsx(xlsx_path)
    
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    sheet_name = _find_sheet(wb)
    ws = wb[sheet_name]

    # Try auto-detection
    detected = _auto_detect_layout(ws)
    
    if detected:
        header_row, field_cols = detected
        first_data_row = _find_first_data_row(ws, header_row, field_cols)
        print(f"      -> Auto-detected header at row {header_row}, data starts at row {first_data_row}")
        print(f"      -> Mapped {len(field_cols)} columns: {list(field_cols.keys())}")

        # Check if this is a pivoted census (Product Type + Plan Name layout)
        from census_normalizer import detect_census_layout, normalize_pivoted_census
        layout = detect_census_layout(field_cols)
        if layout == "pivoted":
            print(f"      -> Detected PIVOTED census layout (Product Type + Plan Name columns)")
            rows = normalize_pivoted_census(ws, header_row, first_data_row, field_cols)
            print(f"      -> Normalized {len(rows)} unique person rows from pivoted data")
            return rows
    else:
        field_cols = CENSUS_FIELD_COLS
        first_data_row = CENSUS_FIRST_DATA_ROW
        print(f"      -> Auto-detect failed, using fallback layout (header row {CENSUS_HEADER_ROW})")

    rows: list[CensusRow] = []
    r = first_data_row
    while True:
        first_name_col = field_cols.get("first_name")
        last_name_col = field_cols.get("last_name")
        
        first_name = ws.cell(row=r, column=first_name_col).value if first_name_col else None
        last_name = ws.cell(row=r, column=last_name_col).value if last_name_col else None
        
        if first_name is None and last_name is None:
            # Two consecutive fully-empty rows -> assume end of table.
            next_first = ws.cell(row=r + 1, column=first_name_col).value if first_name_col else None
            next_last = ws.cell(row=r + 1, column=last_name_col).value if last_name_col else None
            if next_first is None and next_last is None:
                break
            r += 1
            continue
            
        if str(first_name).strip() == "#REF!" or str(last_name).strip() == "#REF!":
            # Skip rows where spreadsheet calculation failed
            r += 1
            continue

        def get(field_name, default=None):
            col = field_cols.get(field_name)
            if col is None:
                return default
            return _clean(ws.cell(row=r, column=col).value)
            
        tier_str = get("tier_generic", "")
        med_tier = get("medical_coverage_tier")
        den_tier = get("dental_coverage_tier")
        vis_tier = get("vision_coverage_tier")
        if tier_str:
            parts = [p.strip() for p in tier_str.split(';')]
            for p in parts:
                if p.startswith('M'):
                    med_tier = p
                elif p.startswith('D'):
                    den_tier = p
                elif p.startswith('V'):
                    vis_tier = p

        rows.append(
            CensusRow(
                row_no=get("row_no"),
                first_name=_clean(first_name),
                last_name=_clean(last_name),
                gender=get("gender"),
                dob=get("dob"),
                home_zip=get("home_zip"),
                relationship=get("relationship"),
                dependent_of_employee_number=get("dependent_of_employee_number"),
                medical_coverage_tier=med_tier,
                medical_plan_enrolled=get("medical_plan_enrolled"),
                medical_plan_price=get("medical_plan_price"),
                cobra=get("cobra"),
                dental_coverage_tier=den_tier,
                dental_plan_enrolled=get("dental_plan_enrolled"),
                dental_plan_price=get("dental_plan_price"),
                vision_coverage_tier=get("vision_coverage_tier"),
                vision_plan_enrolled=get("vision_plan_enrolled"),
                vision_plan_price=get("vision_plan_price"),
                life_plan_name=get("life_plan_name"),
                life_benefit=get("life_benefit"),
                life_rate=get("life_rate"),
                ltd_plan=get("ltd_plan"),
                ltd_benefit=get("ltd_benefit"),
                ltd_rate=get("ltd_rate"),
                std_plan=get("std_plan"),
                std_benefit=get("std_benefit"),
                std_rate=get("std_rate"),
                work_state=get("work_state"),
                job_title=get("job_title"),
                workers_comp_code=get("workers_comp_code"),
                annual_salary=get("annual_salary"),
                ft_pt=get("ft_pt"),
            )
        )
        r += 1

    # -----------------------------------------------------------------------
    # Infer dependent_of_employee_number when the column was not in the census
    # -----------------------------------------------------------------------
    # Guard: only runs when the census has no "Dependent Of" column AND
    # relationship data is present. Censuses with the column are unaffected.
    has_dep_of_col = "dependent_of_employee_number" in field_cols
    has_any_relationship = any(row.relationship for row in rows)
    
    if not has_dep_of_col and has_any_relationship:
        last_ee_row_index = None  # 1-based row_no of the last employee
        for idx, row in enumerate(rows):
            rel_upper = (row.relationship or "").strip().upper()
            if rel_upper in ("EE", "EMPLOYEE", "") or not rel_upper:
                # This is an employee — assign sequential row_no if missing
                if row.row_no is None:
                    row.row_no = idx + 1
                last_ee_row_index = row.row_no
            else:
                # This is a dependent (SP, CH, etc.) — link to last employee
                if row.dependent_of_employee_number is None and last_ee_row_index is not None:
                    row.dependent_of_employee_number = last_ee_row_index

    return rows


def _convert_xls_to_xlsx(xls_path: Path) -> Path:
    """Convert a .xls file to .xlsx format for openpyxl compatibility."""
    try:
        import xlrd
        import openpyxl
        
        xlsx_path = xls_path.with_suffix(".xlsx")
        
        xls_wb = xlrd.open_workbook(str(xls_path))
        xlsx_wb = openpyxl.Workbook()
        
        for sheet_idx, sheet_name in enumerate(xls_wb.sheet_names()):
            xls_ws = xls_wb.sheet_by_name(sheet_name)
            if sheet_idx == 0:
                xlsx_ws = xlsx_wb.active
                xlsx_ws.title = sheet_name
            else:
                xlsx_ws = xlsx_wb.create_sheet(title=sheet_name)
            
            for row_idx in range(xls_ws.nrows):
                for col_idx in range(xls_ws.ncols):
                    cell_value = xls_ws.cell_value(row_idx, col_idx)
                    xlsx_ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
        
        xlsx_wb.save(str(xlsx_path))
        print(f"      -> Converted {xls_path.name} to {xlsx_path.name}")
        return xlsx_path
        
    except ImportError:
        raise ImportError("xlrd is required to read .xls files. Run `pip install xlrd`")


def _convert_csv_to_xlsx(csv_path: Path) -> Path:
    """Convert a .csv file to .xlsx format for openpyxl compatibility."""
    import csv
    import openpyxl
    
    xlsx_path = csv_path.with_suffix(".xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = CENSUS_SHEET_NAME
    
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, start=1):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
                
    wb.save(str(xlsx_path))
    print(f"      -> Converted {csv_path.name} to {xlsx_path.name}")
    return xlsx_path


def save_census_to_excel(rows: list[CensusRow], path: Path) -> None:
    """Save extracted census rows as a structured Excel file."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Census Extracted"
    headers = [
        "Row No", "First Name", "Last Name", "Gender", "DOB",
        "Home Zip", "Relationship", "Dependent Of", "Medical Tier",
        "COBRA", "Dental Tier", "Dental Plan", "Dental Price",
        "Vision Tier", "Vision Plan", "Vision Price",
        "Life Plan", "Life Benefit", "Life Rate",
        "LTD Plan", "LTD Benefit", "LTD Rate",
        "STD Plan", "STD Benefit", "STD Rate",
        "Work State", "Job Title", "Workers Comp", "Annual Salary", "FT/PT",
    ]
    ws.append(headers)
    for row in rows:
        ws.append([
            row.row_no, row.first_name, row.last_name, row.gender,
            row.dob, row.home_zip, row.relationship,
            row.dependent_of_employee_number, row.medical_coverage_tier,
            row.cobra, row.dental_coverage_tier, row.dental_plan_enrolled,
            row.dental_plan_price, row.vision_coverage_tier,
            row.vision_plan_enrolled, row.vision_plan_price,
            row.life_plan_name, row.life_benefit, row.life_rate,
            row.ltd_plan, row.ltd_benefit, row.ltd_rate,
            row.std_plan, row.std_benefit, row.std_rate,
            row.work_state, row.job_title, row.workers_comp_code,
            row.annual_salary, row.ft_pt,
        ])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


def normalize_dob(value) -> object:
    """Return an ISO date string if value looks like a date, else pass through."""
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    return value
