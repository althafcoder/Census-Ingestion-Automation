"""
fill_template.py
----------------
Writes merged census+PDF records into a copy of the Prestige template,
using the exact column layout declared in config.TEMPLATE_COLUMNS.

We open the *template* workbook itself (not a blank one) so all existing
formatting, merged header cells, and the "Census" sheet name are preserved
untouched -- only the data rows are populated.
"""
from __future__ import annotations

import copy
import re
import shutil
from datetime import datetime, date
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string

from config import (
    FIELD_TO_COL,
    TEMPLATE_FIRST_DATA_ROW,
    TEMPLATE_SHEET_NAME,
    DEPENDENT_RELATIONSHIP_MAP,
    MEDICAL_TIER_MAP,
)
from reconcile import MergedRecord


def _col_letter_to_index(letter: str) -> int:
    return column_index_from_string(letter)


def _map_tier(tier: str | None) -> str | None:
    if not tier:
        return None
    raw = str(tier).strip().upper()
    if " - " in raw:
        raw = raw.split(" - ")[0].strip()
    return MEDICAL_TIER_MAP.get(raw, raw)


def _record_to_row_values(record: MergedRecord, row_no: int, last_ee_row_no: int | None = None) -> dict[str, any]:
    row = record.census_row
    summary = record.plan_summary

    # Gender formatting
    gender = (row.gender or "").strip().upper()
    if gender.startswith("F"):
        gender = "F"
    elif gender.startswith("M"):
        gender = "M"

    # Zip code formatting
    zip_val = str(row.home_zip or "").strip()
    if re.match(r'^\d{5}-\d{4}$', zip_val):
        zip_val = zip_val[:5]

    # Cobra default
    cobra = row.cobra or "N"

    # Relationship formatting
    rel = (row.relationship or "").strip().upper()
    relationship = DEPENDENT_RELATIONSHIP_MAP.get(rel, rel)
    
    if row.is_dependent:
        med_tier = _map_tier(row.medical_coverage_tier) or _map_tier(summary.get("medical_coverage_tier"))
        med_waived = med_tier and str(med_tier).strip().upper() in ("WAIVE", "WAIVED", "WO")
        
        den_tier = _map_tier(row.dental_coverage_tier) or _map_tier(summary.get("dental_coverage_tier"))
        den_waived = den_tier and str(den_tier).strip().upper() in ("WAIVE", "WAIVED", "WO")
        
        vis_tier = _map_tier(row.vision_coverage_tier) or _map_tier(summary.get("vision_coverage_tier"))
        vis_waived = vis_tier and str(vis_tier).strip().upper() in ("WAIVE", "WAIVED", "WO")

        return {
            "row_no": row_no,
            "first_name": row.first_name,
            "last_name": row.last_name,
            "gender": gender,
            "dob": row.dob,
            "home_zip": zip_val,
            "relationship": relationship,
            "dependent_of_employee_number": row.dependent_of_employee_number or last_ee_row_no,
            "medical_coverage_tier": med_tier,
            "cobra": cobra,
            "medical_plan_enrolled": None if med_waived else (row.medical_plan_enrolled or summary.get("medical_plan_enrolled")),
            "medical_plan_price": None if med_waived else (summary.get("medical_plan_price") or row.medical_plan_price),
            "dental_coverage_tier": den_tier,
            "dental_plan_enrolled": None if den_waived else (row.dental_plan_enrolled or summary.get("dental_plan_enrolled")),
            "dental_plan_price": None if den_waived else (summary.get("dental_plan_price") or row.dental_plan_price),
            "vision_coverage_tier": vis_tier,
            "vision_plan_enrolled": None if vis_waived else (row.vision_plan_enrolled or summary.get("vision_plan_enrolled")),
            "vision_plan_price": None if vis_waived else (summary.get("vision_plan_price") or row.vision_plan_price),
            "life_plan_name": None,
            "life_benefit": None,
            "life_rate": None,
            "ltd_plan": None,
            "ltd_benefit": None,
            "ltd_rate": None,
            "std_plan": None,
            "std_benefit": None,
            "std_rate": None,
            "work_state": row.work_state,
            "job_title": row.job_title,
            "workers_comp_code": row.workers_comp_code,
            "annual_salary": row.annual_salary,
            "ft_pt": row.ft_pt,
            "discrepancies": "; ".join(record.discrepancies) if record.discrepancies else record.discrepancy_status,
        }

    # For employees
    med_tier = _map_tier(row.medical_coverage_tier) or _map_tier(summary.get("medical_coverage_tier"))
    med_waived = med_tier and str(med_tier).strip().upper() in ("WAIVE", "WAIVED", "WO")
    
    den_tier = _map_tier(row.dental_coverage_tier) or _map_tier(summary.get("dental_coverage_tier"))
    den_waived = den_tier and str(den_tier).strip().upper() in ("WAIVE", "WAIVED", "WO")
    
    vis_tier = _map_tier(row.vision_coverage_tier) or _map_tier(summary.get("vision_coverage_tier"))
    vis_waived = vis_tier and str(vis_tier).strip().upper() in ("WAIVE", "WAIVED", "WO")

    return {
        "row_no": row_no,
        "first_name": row.first_name,
        "last_name": row.last_name,
        "gender": gender,
        "dob": row.dob,
        "home_zip": zip_val,
        "relationship": relationship,
        "dependent_of_employee_number": "",  # blank for employees
        "medical_coverage_tier": med_tier,
        "cobra": cobra,
        "medical_plan_enrolled": None if med_waived else (row.medical_plan_enrolled or summary.get("medical_plan_enrolled")),
        "medical_plan_price": None if med_waived else (summary.get("medical_plan_price") or row.medical_plan_price),
        "dental_coverage_tier": den_tier,
        "dental_plan_enrolled": None if den_waived else (row.dental_plan_enrolled or summary.get("dental_plan_enrolled")),
        "dental_plan_price": None if den_waived else (summary.get("dental_plan_price") or row.dental_plan_price),
        "vision_coverage_tier": vis_tier,
        "vision_plan_enrolled": None if vis_waived else (row.vision_plan_enrolled or summary.get("vision_plan_enrolled")),
        "vision_plan_price": None if vis_waived else (summary.get("vision_plan_price") or row.vision_plan_price),
        "life_plan_name": row.life_plan_name or summary.get("life_plan_name"),
        "life_benefit": row.life_benefit or summary.get("life_benefit"),
        "life_rate": summary.get("life_rate") or row.life_rate,
        "ltd_plan": row.ltd_plan or summary.get("ltd_plan"),
        "ltd_benefit": row.ltd_benefit or summary.get("ltd_benefit"),
        "ltd_rate": summary.get("ltd_rate") or row.ltd_rate,
        "std_plan": row.std_plan or summary.get("std_plan"),
        "std_benefit": row.std_benefit or summary.get("std_benefit"),
        "std_rate": summary.get("std_rate") or row.std_rate,
        "work_state": row.work_state,
        "job_title": row.job_title,
        "workers_comp_code": row.workers_comp_code,
        "annual_salary": row.annual_salary,
        "ft_pt": row.ft_pt,
        "discrepancies": "; ".join(record.discrepancies) if record.discrepancies else record.discrepancy_status,
    }


def fill_template(
    records: list[MergedRecord],
    template_path: Path,
    output_path: Path,
) -> Path:
    """
    Copy `template_path` to `output_path`, then populate the Census sheet's
    data rows (starting at TEMPLATE_FIRST_DATA_ROW) from `records`.
    Copies cell formatting from the first data row to all subsequent rows.
    Also populates the company information header block.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(template_path, output_path)

    wb = openpyxl.load_workbook(str(output_path))
    ws = wb[TEMPLATE_SHEET_NAME]

    # Capture formatting from the first data row (row 26) as a reference
    ref_row = TEMPLATE_FIRST_DATA_ROW
    max_col = max(_col_letter_to_index(col) for _, col in FIELD_TO_COL.items())
    ref_formats = {}
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=ref_row, column=col_idx)
        ref_formats[col_idx] = {
            "font": copy.copy(cell.font),
            "border": copy.copy(cell.border),
            "fill": copy.copy(cell.fill),
            "alignment": copy.copy(cell.alignment),
            "number_format": cell.number_format,
        }

    last_ee_row_no = None
    for i, record in enumerate(records):
        target_row = TEMPLATE_FIRST_DATA_ROW + i
        current_row_no = i + 1

        if not record.census_row.is_dependent:
            last_ee_row_no = current_row_no

        values = _record_to_row_values(record, row_no=current_row_no, last_ee_row_no=last_ee_row_no)

        # Determine if this row needs a specific highlight color in the discrepancies column
        fill_color = None
        status = getattr(record, "discrepancy_status", "")
        method = getattr(record, "match_method", "")
        
        # Color Coding Rules from specification:
        if method == "llm":
            fill_color = "DDEBF7"  # Light Blue
        elif method in ("possible",):
            fill_color = "FFD966"  # Orange
        elif method in ("fuzzy", "initial"):
            fill_color = "FFEB9C"  # Yellow
        elif method in ("exact", "canonical", "token_swap", "token swap"):
            fill_color = "C6EFCE"  # Green
            
        if status in ("Not on census", "mismatch employee name", "mismatch coverage name", "mismatch employee name & mismatch coverage name"):
            fill_color = "FFC7CE"  # Red
        elif status in ("not available on invoice", "Duplicate names listed in the census"):
            fill_color = "FFEB9C"  # Yellow
        elif status == "Correct" and not fill_color:
            fill_color = "C6EFCE"  # Green

        for field_name, value in values.items():
            col_letter = FIELD_TO_COL[field_name]
            col_idx = _col_letter_to_index(col_letter)
            
            # Explicitly clear cell if value is None, as ws.cell(value=None) does not overwrite
            cell = ws.cell(row=target_row, column=col_idx)
            cell.value = value

            # Apply formatting from reference row
            if col_idx in ref_formats:
                fmt = ref_formats[col_idx]
                cell.font = copy.copy(fmt["font"])
                cell.border = copy.copy(fmt["border"])
                
                if fill_color and field_name == "discrepancies":
                    from openpyxl.styles import PatternFill
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                else:
                    cell.fill = copy.copy(fmt["fill"])
                    
                cell.alignment = copy.copy(fmt["alignment"])
                
                # Do not overwrite date format with General for datetime objects
                if isinstance(value, (datetime, date)) and fmt["number_format"] == "General":
                    cell.number_format = "mm/dd/yyyy"
                else:
                    cell.number_format = fmt["number_format"]

    wb.save(str(output_path))
    return output_path
