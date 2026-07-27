"""
census_extractor.py
--------------------
Deterministic reader for the raw "Census.xlsx" demographic workbook
(same layout family as the Prestige template, but without plan/price
columns filled in).

Uses openpyxl directly (rather than pandas.read_excel) because the file
has a multi-row letterhead/header block above the real data table, and
openpyxl lets us address the known header/data rows precisely (see
config.CENSUS_HEADER_ROW / CENSUS_FIRST_DATA_ROW) instead of guessing.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, date
from pathlib import Path

import openpyxl

from config import CENSUS_FIELD_COLS, CENSUS_FIRST_DATA_ROW, CENSUS_SHEET_NAME


@dataclass
class CensusRow:
    row_no: object
    first_name: str
    last_name: str
    gender: str | None
    dob: object
    home_zip: str | None
    relationship: str | None
    dependent_of_employee_number: object
    medical_coverage_tier: str | None
    cobra: str | None
    dental_coverage_tier: str | None = None
    dental_plan_enrolled: str | None = None
    dental_plan_price: object = None
    vision_coverage_tier: str | None = None
    vision_plan_enrolled: str | None = None
    vision_plan_price: object = None
    life_plan_name: str | None = None
    life_benefit: object = None
    life_rate: object = None
    ltd_plan: str | None = None
    ltd_benefit: object = None
    ltd_rate: object = None
    std_plan: str | None = None
    std_benefit: object = None
    std_rate: object = None
    work_state: str | None = None
    job_title: str | None = None
    workers_comp_code: object = None
    annual_salary: object = None
    ft_pt: str | None = None

    @property
    def full_name_key(self) -> str:
        return f"{(self.first_name or '').strip().lower()} {(self.last_name or '').strip().lower()}"

    @property
    def is_dependent(self) -> bool:
        return (self.relationship or "").strip().upper() != "EE"


def _clean(value):
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def extract_census(xlsx_path: Path) -> list[CensusRow]:
    """Read every populated data row from Census.xlsx into CensusRow objects."""
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb[CENSUS_SHEET_NAME]

    rows: list[CensusRow] = []
    r = CENSUS_FIRST_DATA_ROW
    while True:
        first_name = ws.cell(row=r, column=CENSUS_FIELD_COLS["first_name"]).value
        last_name = ws.cell(row=r, column=CENSUS_FIELD_COLS["last_name"]).value
        if first_name is None and last_name is None:
            # Two consecutive fully-empty rows -> assume end of table.
            next_first = ws.cell(row=r + 1, column=CENSUS_FIELD_COLS["first_name"]).value
            next_last = ws.cell(row=r + 1, column=CENSUS_FIELD_COLS["last_name"]).value
            if next_first is None and next_last is None:
                break
            r += 1
            continue

        def get(field_name, default=None):
            col = CENSUS_FIELD_COLS.get(field_name)
            if col is None:
                return default
            return _clean(ws.cell(row=r, column=col).value)

        rows.append(
            CensusRow(
                row_no=get("row_no"),
                first_name=first_name,
                last_name=last_name,
                gender=get("gender"),
                dob=get("dob"),
                home_zip=get("home_zip"),
                relationship=get("relationship"),
                dependent_of_employee_number=get("dependent_of_employee_number"),
                medical_coverage_tier=get("medical_coverage_tier"),
                cobra=get("cobra"),
                dental_coverage_tier=get("dental_coverage_tier"),
                dental_plan_enrolled=get("dental_plan_enrolled"),
                dental_plan_price=get("dental_plan_price"),
                vision_coverage_tier=get("vision_coverage_tier"),
                vision_plan_enrolled=get("vision_plan_enrolled"),
                vision_plan_price=get("vision_plan_price"),
                life_plan_name=get("life_plan_name"),
                life_benefit=get("life_benefit"),
                life_rate=get("life_rate"),
                ltd_plan=get("ltd_plan"),
                ltd_benefit=get("ltd_benefit"),
                ltd_rate=get("ltd_rate"),
                std_plan=get("std_plan"),
                std_benefit=get("std_benefit"),
                std_rate=get("std_rate"),
                work_state=get("work_state"),
                job_title=get("job_title"),
                workers_comp_code=get("workers_comp_code"),
                annual_salary=get("annual_salary"),
                ft_pt=get("ft_pt"),
            )
        )
        r += 1

    return rows


def normalize_dob(value) -> object:
    """Return an ISO date string if value looks like a date, else pass through."""
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    return value
